'''
Created on 8 août 2017

@author: Wentong ZHANG
'''
from __future__ import print_function
import httplib2
import os

from apiclient import discovery
from oauth2client import client
from oauth2client import tools
from oauth2client.file import Storage

import datetime

try:
    import argparse
    flags = argparse.ArgumentParser(parents=[tools.argparser]).parse_args()
except ImportError:
    flags = None

# If modifying these scopes, delete your previously saved credentials
# at ~/.credentials/calendar-python-quickstart.json
SCOPES = 'https://www.googleapis.com/auth/calendar'
CLIENT_SECRET_FILE = 'client_secret.json'
APPLICATION_NAME = 'Google Calendar API Python Quickstart'

def get_credentials():
    home_dir = os.path.expanduser('~')
    credential_dir = os.path.join(home_dir, '.credentials')
    if not os.path.exists(credential_dir):
        os.makedirs(credential_dir)
    credential_path = os.path.join(credential_dir,
                                   'calendar-python-quickstart.json')

    store = Storage(credential_path)
    credentials = store.get()
    if not credentials or credentials.invalid:
        flow = client.flow_from_clientsecrets(CLIENT_SECRET_FILE, SCOPES)
        flow.user_agent = APPLICATION_NAME
        if flags:
            credentials = tools.run_flow(flow, store, flags)
        else: # Needed only for compatibility with Python 2.6
            credentials = tools.run(flow, store)
        print('Storing credentials to ' + credential_path)
    return credentials

def pushlesson(cours,coursetype,description,date,fromtime,totime,resource,groups,learners,instructors):
    # delete all existed events in the google calendar during these 10 weeks
    credentials = get_credentials()
    http = credentials.authorize(httplib2.Http())
    service = discovery.build('calendar', 'v3', http=http)

    # the events of the timetable you choose for next 10 weeks
    now = datetime.datetime.utcnow()
    now -= datetime.timedelta(days=now.weekday())
    now = now.replace(hour=0, minute=0, second=1)
    min = now.isoformat() + 'Z'  # 'Z' indicates UTC time # 2017-08-14T17:12:17.229912Z for 2017.08.14 19h12
    max = (now + datetime.timedelta(days=6, weeks=9)).isoformat() + 'Z'
    # calendarId found by page of calendar which you want to use. default is primary
    eventsResult = service.events().list(
        calendarId='215vrsdu75s3ogbsdh58q8uj8c@group.calendar.google.com', timeMin=min, timeMax=max, singleEvents=True,
        orderBy='startTime').execute()
    events = eventsResult.get('items', [])
    if not events:
        print('No upcoming events found.')
    for event in events:
        service.events().delete(calendarId='215vrsdu75s3ogbsdh58q8uj8c@group.calendar.google.com', eventId=event['id']).execute()

    # start to create and push new events into google calendar
    for i in range(len(date)):
        # transform date form at first
        dates = date[i].split(' ')
        del dates[0]
        if dates[1] == 'janvier':
            dates[1] = '01'
        elif dates[1] == 'février':
            dates[1] = '02'
        elif dates[1] == 'mars':
            dates[1] = '03'
        elif dates[1] == 'avril':
            dates[1] = '04'
        elif dates[1] == 'mai':
            dates[1] = '05'
        elif dates[1] == 'juin':
            dates[1] = '06'
        elif dates[1] == 'juillet':
            dates[1] = '07'
        elif dates[1] == 'août':
            dates[1] = '08'
        elif dates[1] == 'septembre':
            dates[1] = '09'
        elif dates[1] == 'octobre':
            dates[1] = '10'
        elif dates[1] == 'novembre':
            dates[1] = '11'
        elif dates[1] == 'décembre':
            dates[1] = '12'
        newdate = dates[0] + dates[1] + dates[2]
        del dates[:]
        dateok = datetime.datetime.strptime(newdate, "%d%m%Y")
        # date transform end
        # get start time
        start_min = int(fromtime[i].split(':')[1])
        start_hour = int(fromtime[i].split(':')[0])
        start_date = dateok.replace(hour=start_hour, minute=start_min, second=0)
        # get finish time
        end_min = int(totime[i].split(':')[1])
        end_hour = int(totime[i].split(':')[0])
        end_date = dateok.replace(hour=end_hour, minute=end_min, second=0)

        if learners[i].find('/') == -1:
            student = '???'
        else:
            student = learners[i].split('/')[0]

        # start to push data into calendars
        event = {
          'summary': cours[i],
          'location': resource[i],
          'description': coursetype[i] +' - '+description[i],
          'start': {
            'dateTime': start_date.isoformat(),
            'timeZone': 'Europe/Paris',
          },
          'end': {
            'dateTime': end_date.isoformat(),
            'timeZone': 'Europe/Paris',
          },
          'attendees': {
            'comment':  student,
          },
          "organizer": {
            "displayName": instructors[i],
          },
        }
        event = service.events().insert(calendarId='215vrsdu75s3ogbsdh58q8uj8c@group.calendar.google.com', body=event).execute()

def compare(cours,coursetype,description,date,fromtime,totime,resource,groups,learners,instructors):
    from openpyxl import load_workbook
    from openpyxl.styles import Font,colors
    wb = load_workbook(filename='./E_services/Cours.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')
    dest_filename = './E_services/Cours.xlsx'
    found = False # false for totally or included by the old one, true means find the difference 
    print(str(ws.max_row)+' lines in xlsx')
    # check every line of cours, one by one. the most important 3 points is date, start and end time
    if len(date) > ws.max_row:
        found = True
    else:
        for i in range(len(date)):
            if(ws.cell(row=i+1,column=1).value != date[i] or ws.cell(row=i+1,column=2).value != cours[i] or ws.cell(row=i+1,column=3).value != fromtime[i] or
                       ws.cell(row=i+1,column=4).value != totime[i] or ws.cell(row=i+1,column=5).value != coursetype[i] or ws.cell(row=i+1,column=6).value != description[i]
               or ws.cell(row=i+1,column=7).value != resource[i] or ws.cell(row=i+1,column=8).value != groups[i] or ws.cell(row=i+1,column=9).value != learners[i] or
                       ws.cell(row=i + 1, column=10).value != instructors[i]):
                found = True

    if found == False:
        return 'unchanged'
    else:
        data = [['']*(max(ws.max_row,len(date))+1) for i in range(11)]
        wb = load_workbook(filename='./E_services/Cours.xlsx')
        sheet = wb.active
        for col in range(1,len(data)):
            for row in range(1,len(data[1])):
                sheet.cell(row=row, column=col).value = data[col][row]
        wb.save(dest_filename)

        for row in range(1,len(date)+1):
            data[1][row] = date[row-1]
            data[2][row] = cours[row-1]
            data[3][row] = fromtime[row-1]
            data[4][row] = totime[row-1]
            data[5][row] = coursetype[row-1]
            data[6][row] = description[row-1]
            data[7][row] = resource[row-1]
            data[8][row] = groups[row-1]
            data[9][row] = learners[row-1]
            data[10][row] = instructors[row-1]

        wb2 = load_workbook(filename='./E_services/Cours.xlsx')
        shet = wb2.active
        for col in range(1,len(data)):
            for row in range(1,len(data[1])):
                shet.cell(row=row, column=col).value = data[col][row]
        wb2.save(dest_filename)

        return 'changed'
 

def analyse(readFile = False, inners = []):
    from bs4 import BeautifulSoup
    import io
    if not readFile:
        blocks = inners
    else:
        with io.open('cours.txt', 'r', encoding='utf-8') as f:
            blocks = f.read().split("[[[Magic]]]")
        del blocks[-1]
    
    coursetype = []
    description = []
    date = []
    fromtime = []
    totime = []
    resource = [] # the room
    groups = []
    learners = []
    instructors = []
    cours = []
    temp = []
    nom = []
    prenom = []

    print('analyse cours start')

    for block in blocks:
        soup = BeautifulSoup(block, "lxml")
        mainDiv = soup.find_all('div')[1]
        detailDiv = mainDiv.find_all(attrs={"class":"ui-panelgrid-content ui-widget-content ui-grid ui-grid-responsive"})[0].find_all(attrs={"class":"ui-panelgrid-cell ui-grid-col-6"})
        assert 'ui-dialog-content' in mainDiv['class']
        assert 'ui-widget-content' in mainDiv['class']
        
        if detailDiv[7].string:
            description.append(detailDiv[7].string.replace('\n',' '))
        else:
            description.append(' ')

        if detailDiv[5].string:
            coursetype.append(detailDiv[5].string.replace('\n',' '))
        else:
            coursetype.append(' ')
        
        tables = mainDiv.find_all('table')
        
        fromToTable = tables[0]
        resourceTable = tables[1]
        instructorsTable = tables[2]
        learnersTable = tables[3]
        gourpsTable = tables[4]
        courseTable = tables[5]

        courseTds = courseTable.find_all('td')
        cour = ''
        if len(courseTds) > 1:
            for i in range(0, len(courseTds), 2):
                cour = cour + courseTds[i].string + " - " + courseTds[i + 1].string
                if(i != len(courseTds)-2):
                    cour = cour + '/'
            cours.append(cour)
        else:
            cours.append(' ')
        
        resourceTds = resourceTable.find_all('td')
        room = ''
        if len(resourceTds) > 1:
            for i in range(0, len(resourceTds), 2):
                room = room + resourceTds[i].string + " - " + resourceTds[i+1].string
                if(i != len(resourceTds)-2):
                    room = room + '/'
            resource.append(room)
        else:
            resource.append(' ')
            
        teacher = ''
        instructorTds = instructorsTable.find_all('td')
        if len(instructorTds) > 1:
            for i in range(0, len(instructorTds), 2):
                teacher = teacher + instructorTds[i].string + " " + instructorTds[i + 1].string
                if(i != len(instructorTds)-2):
                    teacher = teacher + '/'
            instructors.append(teacher)
        else:
            instructors.append(' ')
            
        student = ''
        learnerTds = learnersTable.find_all('td')
        if len(learnerTds) > 1:
            student = student + str(len(learnerTds)/2) + '/'
            del temp[:]
            del nom[:]
            del prenom[:]
            for i in range(0, len(learnerTds)):
                for item in learnerTds[i].strings:
                    temp.append(item)

            for i in range(len(temp)):
                if (temp[i] == 'Nom'):
                    if(temp[i+1] != 'Prénom'):
                        nom.append(temp[i+1])
                    else:
                        nom.append(' ')
                elif(temp[i] == 'Prénom'):
                    if(temp[i+1] != 'Nom' and i != len(temp)-1):
                        prenom.append(temp[i+1])
                    else:
                        prenom.append(' ')
            for i in range(len(nom)):
                student = student + nom[i] + " " + prenom[i]
                if(i != len(nom)-1):
                    student = student + '/'
            learners.append(student)
        else:
            learners.append(' ')
        
        gp = ''
        groupTds = gourpsTable.find_all('td')
        if len(groupTds) > 0:
            for i in range(0, len(groupTds)):
                gp = gp + groupTds[i].string
                if(i != len(instructorTds)-1):
                    gp = gp + '/'
            groups.append(gp)
        else:
            groups.append(' ')        
        
        date.append(fromToTable.find_all('td')[1].string)
        fromtime.append(fromToTable.find_all('td')[3].string)
        totime.append(fromToTable.find_all('td')[7].string)

    if(len(coursetype)==len(date)):
        print('analyse cours finished')
        print('get '+str(len(date))+' lines this time')
        state = compare(cours,coursetype,description,date,fromtime,totime,resource,groups,learners,instructors)
        if state == 'unchanged':
            return state
        elif state == 'changed':
            print('start push all new cours from localhost to google calendar.')
            try:
                pushlesson(cours, coursetype, description, date, fromtime, totime, resource, groups, learners, instructors)
                print('push ok')
            except:
                print('push to calendar bug')
                state = 'calendar bug'
            return state
        else:
            return 'bug'
    else:
        print('cours sth wrong')
        return 'bug'

if __name__ == '__main__':
    analyse(readFile = True)
