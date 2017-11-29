'''
Created on 8 août 2017

@author: Wentong ZHANG
'''
def compare(nbrAbs,dureeAbs,date,reason,duration,timetable,course,instructor):
    from openpyxl import load_workbook
    from openpyxl.styles import Font,colors
    wb = load_workbook(filename=r'./E_services/Abs.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')
    dest_filename = r'./E_services/Abs.xlsx'
    print(str(ws.max_row)+' lines in xlsx, first line for 2 numbers')
    # we don't need check one by one because abs will only be added
    if(int(ws.cell(row=1,column=2).value) == int(nbrAbs)):
        return 'unchanged'
    else:
        data = [[0]*(ws.max_row+1+int(nbrAbs)-int(ws.cell(row=1,column=2).value)) for i in range(7)]
        for col in range(1,ws.max_column+1):
            for row in range(1,ws.max_row+1):
                data[col][row] = ws.cell(row=row,column=col).value

        for row in range(1,int(nbrAbs)-int(ws.cell(row=1,column=2).value)+1):
            data[1][ws.max_row+row] = date[int(ws.cell(row=1,column=2).value)+row-1]
            data[2][ws.max_row+row] = reason[int(ws.cell(row=1,column=2).value)+row-1]
            data[3][ws.max_row+row] = duration[int(ws.cell(row=1,column=2).value)+row-1]
            data[4][ws.max_row+row] = timetable[int(ws.cell(row=1,column=2).value)+row-1]
            data[5][ws.max_row+row] = course[int(ws.cell(row=1,column=2).value)+row-1]
            data[6][ws.max_row+row] = instructor[int(ws.cell(row=1,column=2).value)+row-1]

        data[2][1] = int(nbrAbs)
        data[1][1] = dureeAbs

        wb = load_workbook(filename=r'./E_services/Abs.xlsx')
        sheet = wb.active
        for col in range(1,len(data)):
            for row in range(1,len(data[1])):
                sheet.cell(row=row, column=col).value = data[col][row]
        wb.save(dest_filename)
        return 'changed'

def analyse(readFile = False, inners = []):
    from bs4 import BeautifulSoup
    import io
    if not readFile:
        blocks = inners
    else:
        with io.open('abs.txt', 'r', encoding='utf-8') as f:
            blocks = f.read().split("[[[Magic]]]")
        del blocks[-1]

    print('analyse abs start')
    date = []
    reason = []
    duration = []
    timetable = []
    course = []
    instructor = []
    note = []

    if (blocks[0] == 'zero abs at this moment'):
        print('analyse abs finished, zero')
        return 'zero'
    else:
        for i in range(0,len(blocks)):
            #first block for the total hour and number of absences
            #the rest for the pages of abs
            soup = BeautifulSoup(blocks[i], "lxml")
            if(i == 0):
                nbrAbs = soup.find(id="form:nbrAbs").string
                dureeAbs = soup.find(id="form:dureeAbs").string
            else:
                tableAbs = soup.find(attrs={"class":"ui-datatable-data ui-widget-content"})
                abs = tableAbs.find_all(role="row")
                for item in abs:
                    lines = item.find_all(role="gridcell")#lines = note of one abscent
                    for line in lines:# line = one row of one abscent
                        for string in line.strings:
                            note.append(string)

                for i in range(0,len(note)):
                    if (note[i] == 'Date'):
                        if(note[i+1] != 'Motif d\'absence'):
                            date.append(note[i+1])
                        else:
                            date.append(' ')
                    elif(note[i] == 'Motif d\'absence'):
                        if(note[i+1] != 'Durée'):
                            reason.append(note[i+1])
                        else:
                            reason.append(' ')
                    elif(note[i] == 'Durée'):
                        if(note[i+1] != 'Horaire'):
                            duration.append(note[i+1])
                        else:
                            duration.append(' ')
                    elif(note[i] == 'Horaire'):
                        if(note[i+1] != 'Cours'):
                            timetable.append(note[i+1])
                        else:
                            timetable.append(' ')
                    elif(note[i] == 'Cours'):
                        if(note[i+1] != 'Intervenant'):
                            course.append(note[i+1])
                        else:
                            course.append(' ')
                    elif(note[i] == 'Intervenant'):
                        if(note[i] != 'Date' and i != len(note)-1):
                            instructor.append(note[i+1])
                        else:
                            instructor.append(' ')
                del note[:]#clear note

        if(len(instructor)==len(course)):
            print('analyse abs finished')
            print('get '+str(len(course))+' lines this time')
            return(compare(nbrAbs,dureeAbs,date,reason,duration,timetable,course,instructor))
        else:
            print('abs sth wrong')
            return 'bug'

if __name__ == '__main__':
    analyse(readFile = True)






