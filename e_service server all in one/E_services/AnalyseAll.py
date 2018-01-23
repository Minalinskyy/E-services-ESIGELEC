'''
Created on 8 août 2017

@author: Wentong ZHANG
'''
from __future__ import print_function
import httplib2
import os
from openpyxl import load_workbook
from openpyxl.styles import Font,colors
from bs4 import BeautifulSoup
import io

def compareedt(cours,coursetype,description,date,fromtime,totime,resource,groups,learners,instructors):
    wb = load_workbook(filename='./E_services/Cours.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')
    dest_filename = './E_services/Cours.xlsx'
    found = False # false for totally or included by the old one, true means find the difference 
    print(str(ws.max_row)+' lines in xlsx')
    # check every line of cours, one by one. the most important 3 points is date, start and end time
    if len(date) > ws.max_row:
        found = True
    else:
        for i in range(len(date)):
            if(ws.cell(row=i+1,column=1).value != date[i] or ws.cell(row=i+1,column=2).value != cours[i] or ws.cell(row=i+1,column=3).value != fromtime[i] or
                       ws.cell(row=i+1,column=4).value != totime[i] or ws.cell(row=i+1,column=5).value != coursetype[i] or ws.cell(row=i+1,column=6).value != description[i]
               or ws.cell(row=i+1,column=7).value != resource[i] or ws.cell(row=i+1,column=8).value != groups[i] or ws.cell(row=i+1,column=9).value != learners[i] or
                       ws.cell(row=i + 1, column=10).value != instructors[i]):
                found = True

    if found == False:
        return 'unchanged'
    else:
        data = [['']*(max(ws.max_row,len(date))+1) for i in range(11)]
        wb = load_workbook(filename='./E_services/Cours.xlsx')
        sheet = wb.active
        for col in range(1,len(data)):
            for row in range(1,len(data[1])):
                sheet.cell(row=row, column=col).value = data[col][row]
        wb.save(dest_filename)

        for row in range(1,len(date)+1):
            data[1][row] = date[row-1]
            data[2][row] = cours[row-1]
            data[3][row] = fromtime[row-1]
            data[4][row] = totime[row-1]
            data[5][row] = coursetype[row-1]
            data[6][row] = description[row-1]
            data[7][row] = resource[row-1]
            data[8][row] = groups[row-1]
            data[9][row] = learners[row-1]
            data[10][row] = instructors[row-1]

        wb2 = load_workbook(filename='./E_services/Cours.xlsx')
        shet = wb2.active
        for col in range(1,len(data)):
            for row in range(1,len(data[1])):
                shet.cell(row=row, column=col).value = data[col][row]
        wb2.save(dest_filename)

        return 'changed'
 
def comparenotes(date,matiere,module,epreuve,note):
    wb = load_workbook(filename='/home/e_service/E_services/Notes.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')
    dest_filename = '/home/e_service/E_services/Notes.xlsx'
    print(str(ws.max_row)+' lines in xlsx')
    # we don't need to check one by one because notes will only be added
    if(ws.max_row == len(date)):
        return 'unchanged'
    else:
        data = [[0]*(len(date)+1) for i in range(6)]

        for row in range(1,len(date)+1):
            data[1][row] = date[row-1]
            data[2][row] = matiere[row-1]
            data[3][row] = module[row-1]
            data[4][row] = epreuve[row-1]
            data[5][row] = note[row-1]

        wb = load_workbook(filename='/home/e_service/E_services/Notes.xlsx')
        sheet = wb.active
        for col in range(1,len(data)):
            for row in range(1,len(data[1])):
                sheet.cell(row=row, column=col).value = data[col][row]
        wb.save(dest_filename)
        return 'changed'

def compareabs(nbrAbs,dureeAbs,date,reason,duration,timetable,course,instructor):
    wb = load_workbook(filename='/home/e_service/E_services/Abs.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')
    dest_filename = '/home/e_service/E_services/Abs.xlsx'
    print(str(ws.max_row)+' lines in xlsx, first line for 2 numbers')
    # we don't need check one by one because abs will only be added
    if(int(ws.cell(row=1,column=2).value) == int(nbrAbs)):
        return 'unchanged'
    else:
        data = [[0]*(ws.max_row+1+int(nbrAbs)-int(ws.cell(row=1,column=2).value)) for i in range(7)]
        for col in range(1,ws.max_column+1):
            for row in range(1,ws.max_row+1):
                data[col][row] = ws.cell(row=row,column=col).value

        for row in range(1,int(nbrAbs)-int(ws.cell(row=1,column=2).value)+1):
            data[1][ws.max_row+row] = date[int(ws.cell(row=1,column=2).value)+row-1]
            data[2][ws.max_row+row] = reason[int(ws.cell(row=1,column=2).value)+row-1]
            data[3][ws.max_row+row] = duration[int(ws.cell(row=1,column=2).value)+row-1]
            data[4][ws.max_row+row] = timetable[int(ws.cell(row=1,column=2).value)+row-1]
            data[5][ws.max_row+row] = course[int(ws.cell(row=1,column=2).value)+row-1]
            data[6][ws.max_row+row] = instructor[int(ws.cell(row=1,column=2).value)+row-1]

        data[2][1] = int(nbrAbs)
        data[1][1] = dureeAbs

        wb = load_workbook(filename='/home/e_service/E_services/Abs.xlsx')
        sheet = wb.active
        for col in range(1,len(data)):
            for row in range(1,len(data[1])):
                sheet.cell(row=row, column=col).value = data[col][row]
        wb.save(dest_filename)
        return 'changed'

def analyse(readFile = False, inners = []):
    result = []

    if not readFile:
        blocks = inners[0]
    else:
        with io.open('cours.txt', 'r', encoding='utf-8') as f:
            blocks = f.read().split("[[[Magic]]]")
        del blocks[-1]
    
    coursetype = []
    description = []
    date = []
    fromtime = []
    totime = []
    resource = [] # the room
    groups = []
    learners = []
    instructors = []
    cours = []
    temp = []
    nom = []
    prenom = []

    print('analyse cours start')

    for block in blocks:
        soup = BeautifulSoup(block, "lxml")
        mainDiv = soup.find_all('div')[1]
        detailDiv = mainDiv.find_all(attrs={"class":"ui-panelgrid-content ui-widget-content ui-grid ui-grid-responsive"})[0].find_all(attrs={"class":"ui-panelgrid-cell ui-grid-col-6"})
        assert 'ui-dialog-content' in mainDiv['class']
        assert 'ui-widget-content' in mainDiv['class']
        
        if detailDiv[7].string:
            description.append(detailDiv[7].string.replace('\n',' '))
        else:
            description.append(' ')

        if detailDiv[5].string:
            coursetype.append(detailDiv[5].string.replace('\n',' '))
        else:
            coursetype.append(' ')
        
        tables = mainDiv.find_all('table')
        
        fromToTable = tables[0]
        resourceTable = tables[1]
        instructorsTable = tables[2]
        learnersTable = tables[3]
        gourpsTable = tables[4]
        courseTable = tables[5]

        courseTds = courseTable.find_all('td')
        cour = ''
        if len(courseTds) > 1:
            for i in range(0, len(courseTds), 2):
                cour = cour + courseTds[i].string + " - " + courseTds[i + 1].string
                if(i != len(courseTds)-2):
                    cour = cour + '/'
            cours.append(cour)
        else:
            cours.append(' ')
        
        resourceTds = resourceTable.find_all('td')
        room = ''
        if len(resourceTds) > 1:
            for i in range(0, len(resourceTds), 2):
                room = room + resourceTds[i].string + " - " + resourceTds[i+1].string
                if(i != len(resourceTds)-2):
                    room = room + '/'
            resource.append(room)
        else:
            resource.append(' ')
            
        teacher = ''
        instructorTds = instructorsTable.find_all('td')
        if len(instructorTds) > 1:
            for i in range(0, len(instructorTds), 2):
                teacher = teacher + instructorTds[i].string + " " + instructorTds[i + 1].string
                if(i != len(instructorTds)-2):
                    teacher = teacher + '/'
            instructors.append(teacher)
        else:
            instructors.append(' ')
            
        student = ''
        learnerTds = learnersTable.find_all('td')
        if len(learnerTds) > 1:
            student = student + str(len(learnerTds)/2) + '/'
            del temp[:]
            del nom[:]
            del prenom[:]
            for i in range(0, len(learnerTds)):
                for item in learnerTds[i].strings:
                    temp.append(item)

            for i in range(len(temp)):
                if (temp[i] == 'Nom'):
                    if(temp[i+1] != 'Prénom'):
                        nom.append(temp[i+1])
                    else:
                        nom.append(' ')
                elif(temp[i] == 'Prénom'):
                    if(temp[i+1] != 'Nom' and i != len(temp)-1):
                        prenom.append(temp[i+1])
                    else:
                        prenom.append(' ')
            for i in range(len(nom)):
                student = student + nom[i] + " " + prenom[i]
                if(i != len(nom)-1):
                    student = student + '/'
            learners.append(student)
        else:
            learners.append(' ')
        
        gp = ''
        groupTds = gourpsTable.find_all('td')
        if len(groupTds) > 0:
            for i in range(0, len(groupTds)):
                gp = gp + groupTds[i].string
                if(i != len(instructorTds)-1):
                    gp = gp + '/'
            groups.append(gp)
        else:
            groups.append(' ')        
        
        date.append(fromToTable.find_all('td')[1].string)
        fromtime.append(fromToTable.find_all('td')[3].string)
        totime.append(fromToTable.find_all('td')[7].string)

    if(len(coursetype)==len(date)):
        print('analyse cours finished')
        print('get '+str(len(date))+' lines this time')
        state = compareedt(cours,coursetype,description,date,fromtime,totime,resource,groups,learners,instructors)
        if state == 'unchanged':
            result.append(state)
        elif state == 'changed':
            result.append(state)
        else:
            result.append('bug')
    else:
        print('cours sth wrong')
        result.append('bug')

    if not readFile:
        blocks = inners[1]
    else:
        with io.open('notes.txt', 'r', encoding='utf-8') as f:
            blocks = f.read().split("[[[Magic]]]")
        del blocks[-1]

    print('analyse notes start')
    date = []
    matiere = []
    module = []
    epreuve = []
    note = []
    temp = []

    for i in range(0,len(blocks)):
        soup = BeautifulSoup(blocks[i], "lxml")
        tableNotes = soup.find(attrs={"class":"ui-datatable-data ui-widget-content"})
        notes = tableNotes.find_all(role="row")
        for item in notes:
            lines = item.find_all(role="gridcell")#lines = note of one abscent
            for line in lines:# line = one row of one abscent
                for string in line.strings:
                    temp.append(string)

        for i in range(0,len(temp)):
            if (temp[i] == 'Date de l\'épreuve'):
                if(temp[i+1] != 'Matière'):
                    date.append(temp[i+1])
                else:
                    date.append(' ')
            elif(temp[i] == 'Matière'):
                if(temp[i+1] != 'Module'):
                    matiere.append(temp[i+1])
                else:
                    matiere.append(' ')
            elif(temp[i] == 'Module'):
                if(temp[i+1] != 'Epreuve'):
                    module.append(temp[i+1])
                else:
                    module.append(' ')
            elif(temp[i] == 'Epreuve'):
                if(temp[i+1] != 'Note obtenue'):
                    epreuve.append(temp[i+1])
                else:
                    epreuve.append(' ')
            elif(temp[i] == 'Note obtenue'):
                if(temp[i] != 'Date de l\'épreuve' and i != len(temp)-1):
                    note.append(temp[i+1])
                else:
                    note.append(' ')
        del temp[:]#clear temp

    if(len(note)==len(date)):
        print('analyse notes finished') 
        print('get '+str(len(note))+' lines this time')  
        result.append(comparenotes(date,matiere,module,epreuve,note))
    else:
        print('notes sth wrong')
        result.append('bug')

    if not readFile:
        blocks = inners[2]
    else:
        with io.open('abs.txt', 'r', encoding='utf-8') as f:
            blocks = f.read().split("[[[Magic]]]")
        del blocks[-1]

    print('analyse abs start')
    date = []
    reason = []
    duration = []
    timetable = []
    course = []
    instructor = []
    note = []

    if (blocks[0] == 'zero abs at this moment'):
        print('analyse abs finished, zero')
        result.append('zero')
    else:
        for i in range(0,len(blocks)):
            #first block for the total hour and number of absences
            #the rest for the pages of abs
            soup = BeautifulSoup(blocks[i], "lxml")
            if(i == 0):
                nbrAbs = soup.find(id="form:nbrAbs").string
                dureeAbs = soup.find(id="form:dureeAbs").string
            else:
                tableAbs = soup.find(attrs={"class":"ui-datatable-data ui-widget-content"})
                abs = tableAbs.find_all(role="row")
                for item in abs:
                    lines = item.find_all(role="gridcell")#lines = note of one abscent
                    for line in lines:# line = one row of one abscent
                        for string in line.strings:
                            note.append(string)

                for i in range(0,len(note)):
                    if (note[i] == 'Date'):
                        if(note[i+1] != 'Motif d\'absence'):
                            date.append(note[i+1])
                        else:
                            date.append(' ')
                    elif(note[i] == 'Motif d\'absence'):
                        if(note[i+1] != 'Durée'):
                            reason.append(note[i+1])
                        else:
                            reason.append(' ')
                    elif(note[i] == 'Durée'):
                        if(note[i+1] != 'Horaire'):
                            duration.append(note[i+1])
                        else:
                            duration.append(' ')
                    elif(note[i] == 'Horaire'):
                        if(note[i+1] != 'Cours'):
                            timetable.append(note[i+1])
                        else:
                            timetable.append(' ')
                    elif(note[i] == 'Cours'):
                        if(note[i+1] != 'Intervenant'):
                            course.append(note[i+1])
                        else:
                            course.append(' ')
                    elif(note[i] == 'Intervenant'):
                        if(note[i] != 'Date' and i != len(note)-1):
                            instructor.append(note[i+1])
                        else:
                            instructor.append(' ')
                del note[:]#clear note

        if(len(instructor)==len(course)):
            print('analyse abs finished')
            print('get '+str(len(course))+' lines this time')
            result.append(compareabs(nbrAbs,dureeAbs,date,reason,duration,timetable,course,instructor))
        else:
            print('abs sth wrong')
            result.append('bug')

    return result

if __name__ == '__main__':
    analyse(readFile = True)
