'''
Created on 8 août 2017

@author: Wentong ZHANG
'''

def compare(date,matiere,module,epreuve,note):
    from openpyxl import load_workbook
    from openpyxl.styles import Font,colors
    wb = load_workbook(filename='/home/e_service/E_services/Notes.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')
    dest_filename = '/home/e_service/E_services/Notes.xlsx'
    print(str(ws.max_row)+' lines in xlsx')
    # we don't need to check one by one because notes will only be added
    if(ws.max_row == len(date)):
        return 'unchanged'
    else:
        data = [[0]*(len(date)+1) for i in range(6)]

        for row in range(1,len(date)+1):
            data[1][row] = date[row-1]
            data[2][row] = matiere[row-1]
            data[3][row] = module[row-1]
            data[4][row] = epreuve[row-1]
            data[5][row] = note[row-1]

        wb = load_workbook(filename='/home/e_service/E_services/Notes.xlsx')
        sheet = wb.active
        for col in range(1,len(data)):
            for row in range(1,len(data[1])):
                sheet.cell(row=row, column=col).value = data[col][row]
        wb.save(dest_filename)
        return 'changed'

def analyse(readFile = False, inners = []):
    from bs4 import BeautifulSoup
    import io
    if not readFile:
        blocks = inners
    else:
        with io.open('notes.txt', 'r', encoding='utf-8') as f:
            blocks = f.read().split("[[[Magic]]]")
        del blocks[-1]

    print('analyse notes start')
    date = []
    matiere = []
    module = []
    epreuve = []
    note = []
    temp = []

    for i in range(0,len(blocks)):
        soup = BeautifulSoup(blocks[i], "lxml")
        tableNotes = soup.find(attrs={"class":"ui-datatable-data ui-widget-content"})
        notes = tableNotes.find_all(role="row")
        for item in notes:
            lines = item.find_all(role="gridcell")#lines = note of one abscent
            for line in lines:# line = one row of one abscent
                for string in line.strings:
                    temp.append(string)

        for i in range(0,len(temp)):
            if (temp[i] == 'Date de l\'épreuve'):
                if(temp[i+1] != 'Matière'):
                    date.append(temp[i+1])
                else:
                    date.append(' ')
            elif(temp[i] == 'Matière'):
                if(temp[i+1] != 'Module'):
                    matiere.append(temp[i+1])
                else:
                    matiere.append(' ')
            elif(temp[i] == 'Module'):
                if(temp[i+1] != 'Epreuve'):
                    module.append(temp[i+1])
                else:
                    module.append(' ')
            elif(temp[i] == 'Epreuve'):
                if(temp[i+1] != 'Note obtenue'):
                    epreuve.append(temp[i+1])
                else:
                    epreuve.append(' ')
            elif(temp[i] == 'Note obtenue'):
                if(temp[i] != 'Date de l\'épreuve' and i != len(temp)-1):
                    note.append(temp[i+1])
                else:
                    note.append(' ')
        del temp[:]#clear temp

    if(len(note)==len(date)):
        print('analyse notes finished') 
        print('get '+str(len(note))+' lines this time')  
        return(compare(date,matiere,module,epreuve,note))
    else:
        print('notes sth wrong')
        return 'bug'

if __name__ == '__main__':
    analyse(readFile = True)






