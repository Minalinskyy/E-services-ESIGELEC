'''
Created on 29 novembre 2017

@author: Wentong ZHANG
'''

from openpyxl import load_workbook

def transferdate(date):
    from datetime import datetime
    splited = date.split(' ')
    newStr = splited[1] + '/' + splited[2] + '/' + splited[3]
    newStr = newStr.replace("janvier", "01") \
        .replace("février", "02") \
        .replace("mars", "03") \
        .replace("avril", "04") \
        .replace("mai", "05") \
        .replace("juin", "06") \
        .replace("juillet", "07") \
        .replace("août", "08") \
        .replace("septembre", "09") \
        .replace("octobre", "10") \
        .replace("novembre", "11") \
        .replace("décembre", "12")

    return datetime.strptime(newStr, '%d/%m/%Y')

def readcours():
    lessons = []
    lesson = []

    wb = load_workbook(filename='./E_services/Cours.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')

    for i in range(1, ws.max_row+1):
        if ws.cell(row=i,column=1).value != None:
            lesson.append(transferdate(ws.cell(row=i,column=1).value)) # date is changed into type: datetime like: 2017-11-27 00:00:00
            lesson.append(ws.cell(row=i,column=2).value)
            lesson.append(ws.cell(row=i,column=3).value)
            lesson.append(ws.cell(row=i,column=4).value)
            lesson.append(ws.cell(row=i,column=5).value)
            lesson.append(ws.cell(row=i,column=6).value)
            lesson.append(ws.cell(row=i,column=7).value)
            lesson.append(ws.cell(row=i,column=8).value)
            lesson.append(ws.cell(row=i,column=9).value.split('/')[0])
            lesson.append(ws.cell(row=i,column=10).value)
            # by order: date, cours, fromtime, totime, coursetype, description, resources, groups, learners, instructors
            
            # save this lesson into lessons
            lessons.append(lesson.copy())
            # clear the list of lesson
            del lesson[:]
        else:
            break

    return lessons # return all lessons

def generateICS(lessons):
    from icalendar import Calendar, Event
    from datetime import datetime
    import random
    random.seed(233333)

    cal = Calendar()
    cal.add('prodid', '-//Catprogrammer.com//ESIGELEC Planning//FR')
    cal.add('version', '2.0')
    cal.add('method', 'publish')
    cal.add('X-WR-TIMEZONE', 'Europe/Paris')

    for lesson in lessons:
        e = Event()
        # by order: date, cours, fromtime, totime, coursetype, description, resources,groups, learners, instructors
        #             0     1        2         3          4        5           6        7          8        9
        summary = ''
        if len(lesson[4]) != 0:
            summary += lesson[4]
        summary += ' '
        if len(lesson[1]) != 0:
            summary += lesson[1]
        e.add('summary', summary)

        y = lesson[0].year
        m = lesson[0].month
        d = lesson[0].day
        fromH = int(lesson[2].split(":")[0])
        fromM = int(lesson[2].split(":")[1])
        toH = int(lesson[3].split(":")[0])
        toM = int(lesson[3].split(":")[1])

        e.add('dtstart', datetime(y,m,d,fromH,fromM,0))
        e.add('dtend', datetime(y,m,d,toH,toM,0))
        e.add('dtstamp', datetime.now())
        e.add('uid', lesson[0].strftime('%Y%m%d') + lesson[2].replace(":", "") + '/' + str(random.randint(0, 1000001)) + '@catprogrammer.com')
        description = lesson[5]
        if len(description) != 0:
            description = description.replace('/',',')
        else:
            description = 'unknown'
        e.add('description', lesson[9][0] + ", " + lesson[8] + " students, " + description)

        room = lesson[6]
        if len(room) != 0:
            room = room.replace('/',',')
        else:
            room = 'unknown'
        e.add('location',room)

        instructor = lesson[9]
        if len(instructor) != 0 :
            instructor = instructor.replace('/', ',')
        else:
            instructor = 'unknown'
        e.add('organizer',instructor)

        group = lesson[7]
        if len(group) != 0:
            for item in group.split('/'):
                e.add('attendee',item)

        cal.add_component(e)

    import io
    with io.open('cal.ics','wb') as f:
        f.write(cal.to_ical())


if __name__=="__main__":
    generateICS(readcours())