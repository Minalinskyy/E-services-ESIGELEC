'''
Created on 8 août 2017

@author: Wentong ZHANG
'''
from __future__ import print_function
import httplib2
import os

import datetime

def compare(cours,coursetype,description,date,fromtime,totime,resource,groups,learners,instructors):
    from openpyxl import load_workbook
    from openpyxl.styles import Font,colors
    wb = load_workbook(filename='./E_services/Cours.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')
    dest_filename = './E_services/Cours.xlsx'
    found = False # false for totally or included by the old one, true means find the difference 
    print(str(ws.max_row)+' lines in xlsx')
    # check every line of cours, one by one. the most important 3 points is date, start and end time
    if len(date) > ws.max_row:
        found = True
    else:
        for i in range(len(date)):
            if(ws.cell(row=i+1,column=1).value != date[i] or ws.cell(row=i+1,column=2).value != cours[i] or ws.cell(row=i+1,column=3).value != fromtime[i] or
                       ws.cell(row=i+1,column=4).value != totime[i] or ws.cell(row=i+1,column=5).value != coursetype[i] or ws.cell(row=i+1,column=6).value != description[i]
               or ws.cell(row=i+1,column=7).value != resource[i] or ws.cell(row=i+1,column=8).value != groups[i] or ws.cell(row=i+1,column=9).value != learners[i] or
                       ws.cell(row=i + 1, column=10).value != instructors[i]):
                found = True

    if found == False:
        return 'unchanged'
    else:
        data = [['']*(max(ws.max_row,len(date))+1) for i in range(11)]
        wb = load_workbook(filename='./E_services/Cours.xlsx')
        sheet = wb.active
        for col in range(1,len(data)):
            for row in range(1,len(data[1])):
                sheet.cell(row=row, column=col).value = data[col][row]
        wb.save(dest_filename)

        for row in range(1,len(date)+1):
            data[1][row] = date[row-1]
            data[2][row] = cours[row-1]
            data[3][row] = fromtime[row-1]
            data[4][row] = totime[row-1]
            data[5][row] = coursetype[row-1]
            data[6][row] = description[row-1]
            data[7][row] = resource[row-1]
            data[8][row] = groups[row-1]
            data[9][row] = learners[row-1]
            data[10][row] = instructors[row-1]

        wb2 = load_workbook(filename='./E_services/Cours.xlsx')
        shet = wb2.active
        for col in range(1,len(data)):
            for row in range(1,len(data[1])):
                shet.cell(row=row, column=col).value = data[col][row]
        wb2.save(dest_filename)

        return 'changed'
 

def analyse(readFile = False, inners = []):
    from bs4 import BeautifulSoup
    import io
    if not readFile:
        blocks = inners
    else:
        with io.open('cours.txt', 'r', encoding='utf-8') as f:
            blocks = f.read().split("[[[Magic]]]")
        del blocks[-1]
    
    coursetype = []
    description = []
    date = []
    fromtime = []
    totime = []
    resource = [] # the room
    groups = []
    learners = []
    instructors = []
    cours = []
    temp = []
    nom = []
    prenom = []

    print('analyse cours start')

    for block in blocks:
        soup = BeautifulSoup(block, "lxml")
        mainDiv = soup.find_all('div')[1]
        detailDiv = mainDiv.find_all(attrs={"class":"ui-panelgrid-content ui-widget-content ui-grid ui-grid-responsive"})[0].find_all(attrs={"class":"ui-panelgrid-cell ui-grid-col-6"})
        assert 'ui-dialog-content' in mainDiv['class']
        assert 'ui-widget-content' in mainDiv['class']
        
        if detailDiv[7].string:
            description.append(detailDiv[7].string.replace('\n',' '))
        else:
            description.append(' ')

        if detailDiv[5].string:
            coursetype.append(detailDiv[5].string.replace('\n',' '))
        else:
            coursetype.append(' ')
        
        tables = mainDiv.find_all('table')
        
        fromToTable = tables[0]
        resourceTable = tables[1]
        instructorsTable = tables[2]
        learnersTable = tables[3]
        gourpsTable = tables[4]
        courseTable = tables[5]

        courseTds = courseTable.find_all('td')
        cour = ''
        if len(courseTds) > 1:
            for i in range(0, len(courseTds), 2):
                cour = cour + courseTds[i].string + " - " + courseTds[i + 1].string
                if(i != len(courseTds)-2):
                    cour = cour + '/'
            cours.append(cour)
        else:
            cours.append(' ')
        
        resourceTds = resourceTable.find_all('td')
        room = ''
        if len(resourceTds) > 1:
            for i in range(0, len(resourceTds), 2):
                room = room + resourceTds[i].string + " - " + resourceTds[i+1].string
                if(i != len(resourceTds)-2):
                    room = room + '/'
            resource.append(room)
        else:
            resource.append(' ')
            
        teacher = ''
        instructorTds = instructorsTable.find_all('td')
        if len(instructorTds) > 1:
            for i in range(0, len(instructorTds), 2):
                teacher = teacher + instructorTds[i].string + " " + instructorTds[i + 1].string
                if(i != len(instructorTds)-2):
                    teacher = teacher + '/'
            instructors.append(teacher)
        else:
            instructors.append(' ')
            
        student = ''
        learnerTds = learnersTable.find_all('td')
        if len(learnerTds) > 1:
            student = student + str(len(learnerTds)/2) + '/'
            del temp[:]
            del nom[:]
            del prenom[:]
            for i in range(0, len(learnerTds)):
                for item in learnerTds[i].strings:
                    temp.append(item)

            for i in range(len(temp)):
                if (temp[i] == 'Nom'):
                    if(temp[i+1] != 'Prénom'):
                        nom.append(temp[i+1])
                    else:
                        nom.append(' ')
                elif(temp[i] == 'Prénom'):
                    if(temp[i+1] != 'Nom' and i != len(temp)-1):
                        prenom.append(temp[i+1])
                    else:
                        prenom.append(' ')
            for i in range(len(nom)):
                student = student + nom[i] + " " + prenom[i]
                if(i != len(nom)-1):
                    student = student + '/'
            learners.append(student)
        else:
            learners.append(' ')
        
        gp = ''
        groupTds = gourpsTable.find_all('td')
        if len(groupTds) > 0:
            for i in range(0, len(groupTds)):
                gp = gp + groupTds[i].string
                if(i != len(instructorTds)-1):
                    gp = gp + '/'
            groups.append(gp)
        else:
            groups.append(' ')        
        
        date.append(fromToTable.find_all('td')[1].string)
        fromtime.append(fromToTable.find_all('td')[3].string)
        totime.append(fromToTable.find_all('td')[7].string)

    if(len(coursetype)==len(date)):
        print('analyse cours finished')
        print('get '+str(len(date))+' lines this time')
        state = compare(cours,coursetype,description,date,fromtime,totime,resource,groups,learners,instructors)
        if state == 'unchanged':
            return state
        elif state == 'changed':
            return state
        else:
            return 'bug'
    else:
        print('cours sth wrong')
        return 'bug'

if __name__ == '__main__':
    analyse(readFile = True)
