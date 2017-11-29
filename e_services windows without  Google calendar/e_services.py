'''
Created on 8 août 2017

@author: Wentong ZHANG
'''
from E_services import abs_download, AnalyseAbs, notes_download, AnalyseNotes, cours_download, AnalyseCours
import tkinter as tk
import tkinter.scrolledtext as tkst
from openpyxl import load_workbook
from openpyxl.styles import Font,colors
import tkinter.font as font

# defaul value of username and password
usr = 'null'
pwd = 'null'
choice = 0

def searchcours(text,frame):
    frame.information.delete("1.0",tk.END)
    wb = load_workbook(filename=r'./E_services/Cours.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')

    coursetype = []
    description = []
    date = []
    fromtime = []
    totime = []
    resource = [] # the room
    groups = []
    learners = []
    instructors = []
    cours = []
    found = False

    for i in range(1, ws.max_row+1):
        for j in range(1,11):
            if ws.cell(row=i,column=j).value != None:
                if ws.cell(row=i,column=j).value.find(text) != -1:
                    found = True
                    date.append(ws.cell(row=i,column=1).value)
                    cours.append(ws.cell(row=i,column=2).value)
                    fromtime.append(ws.cell(row=i,column=3).value)
                    totime.append(ws.cell(row=i,column=4).value)
                    coursetype.append(ws.cell(row=i,column=5).value)
                    description.append(ws.cell(row=i,column=6).value)
                    resource.append(ws.cell(row=i,column=7).value)
                    groups.append(ws.cell(row=i,column=8).value)
                    learners.append(ws.cell(row=i,column=9).value)
                    instructors.append(ws.cell(row=i,column=10).value)
                    break
            else:
            	break

    if found == True:
        for i in range(len(date)):
            result = ''
            result += 'Cours: '+cours[i]+'\n'
            result += 'Date: '+date[i]+'\n'
            result += 'Time: from '+fromtime[i]+' to '+totime[i]+'\n'
            result += 'Room: '+resource[i]+'\n'
            if learners[i].find('/') == -1:
                result += 'Students: ??? \n'
            else:
                result += 'Students: '+learners[i].split('/')[0]+'\n'
            if groups[i].find('/') == -1:
                result += 'Groups: ???\n'
            else:
                for j in range(len(groups[i].split('/'))):
                    if j == 0:
                        result += 'Groups: '
                    result += groups[i].split('/')[j]+'\n'
            result += 'Instructors: '+instructors[i]+'\n'
            result += 'Type and description: '+coursetype[i] +' - '+description[i]+'\n'
            result += '-'*150
            result += '\n\n\n'
            frame.information.insert(tk.INSERT,result)
    else:
        result = 'No corresponding item found.\n'
        frame.information.insert(tk.INSERT,result)

def readcours(frame):
    coursetype = []
    description = []
    date = []
    fromtime = []
    totime = []
    resource = [] # the room
    groups = []
    learners = []
    instructors = []
    cours = []

    frame.information.delete("1.0",tk.END)
    wb = load_workbook(filename=r'./E_services/Cours.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')

    for i in range(1, ws.max_row+1):
        if ws.cell(row=i,column=1).value != None:
            date.append(ws.cell(row=i,column=1).value)
            cours.append(ws.cell(row=i,column=2).value)
            fromtime.append(ws.cell(row=i,column=3).value)
            totime.append(ws.cell(row=i,column=4).value)
            coursetype.append(ws.cell(row=i,column=5).value)
            description.append(ws.cell(row=i,column=6).value)
            resource.append(ws.cell(row=i,column=7).value)
            groups.append(ws.cell(row=i,column=8).value)
            learners.append(ws.cell(row=i,column=9).value)
            instructors.append(ws.cell(row=i,column=10).value)
        else:
            break

    for i in range(len(date)):
        result = ''
        result += 'Cours: '+cours[i]+'\n'
        result += 'Date: '+date[i]+'\n'
        result += 'Time: from '+fromtime[i]+' to '+totime[i]+'\n'
        result += 'Room: '+resource[i]+'\n'
        if learners[i].find('/') == -1:
            result += 'Students: ??? \n'
        else:
            result += 'Students: '+learners[i].split('/')[0]+'\n'
        if groups[i].find('/') == -1:
            result += 'Groups: ???\n'
        else:
            for j in range(len(groups[i].split('/'))):
                if j == 0:
                    result += 'Groups: '
                result += groups[i].split('/')[j]+'\n'
        result += 'Instructors: '+instructors[i]+'\n'
        result += 'Type and description: '+coursetype[i] +' - '+description[i]+'\n'
        result += '-'*150
        result += '\n\n\n'
        frame.information.insert(tk.INSERT,result)

def readabs(frame):
    date = []
    reason = []
    duration = []
    timetable = []
    course = []
    instructor = []

    frame.information.delete("1.0",tk.END)
    wb = load_workbook(filename=r'./E_services/Abs.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')

    frame.information.insert(tk.INSERT,'Abs number: '+str(ws.cell(row=1,column=2).value)+'  Abs time: '+ws.cell(row=1,column=1).value+'\n\n\n')

    for i in range(2, ws.max_row+1):
        date.append(ws.cell(row=i,column=1).value)
        reason.append(ws.cell(row=i,column=2).value)
        duration.append(ws.cell(row=i,column=3).value)
        timetable.append(ws.cell(row=i,column=4).value)
        course.append(ws.cell(row=i,column=5).value)
        instructor.append(ws.cell(row=i,column=6).value)

    for i in range(len(date)):
        result = ''
        result += 'Date: '+date[i]+'\n'
        result += 'Timetable: '+timetable[i]+'\n'
        result += 'Duration: '+duration[i]+'\n'
        result += 'Cours: '+course[i]+'\n'
        result += 'Instructor: '+instructor[i]+'\n'
        result += 'Reason: '+reason[i]+'\n'
        result += '-'*150
        result += '\n\n'
        frame.information.insert(tk.INSERT,result)

def readnotes(frame):
    date = []
    matiere = []
    module = []
    epreuve = []
    note = []

    frame.information.delete("1.0",tk.END)
    wb = load_workbook(filename=r'./E_services/Notes.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')

    for i in range(1, ws.max_row+1):
        date.append(ws.cell(row=i,column=1).value)
        matiere.append(ws.cell(row=i,column=2).value)
        module.append(ws.cell(row=i,column=3).value)
        epreuve.append(ws.cell(row=i,column=4).value)
        note.append(ws.cell(row=i,column=5).value)

    for i in range(len(date)):
        result = ''
        result += 'Date: '+date[i]+'\n'
        result += 'Matiere: '+matiere[i]+'\n'
        result += 'Module: '+module[i]+'\n'
        result += 'Epreuve: '+epreuve[i]+'\n'
        result += 'Note: '+note[i]+'\n'
        result += '-'*150
        result += '\n\n'
        frame.information.insert(tk.INSERT,result)

class Mainframe(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        self.frame = MainFrame(self)
        self.frame.pack()

    def change(self, frame):
        self.frame.pack_forget() # delete currrent frame
        self.frame = frame(self)
        self.frame.pack() # make new frame

class MainFrame(tk.Frame):
    def __init__(self, master=None, **kwargs):
        global usr
        global pwd
        global choice
        tk.Frame.__init__(self, master, **kwargs)
        master.title("Window to check information")
        master.geometry('%dx%d+%d+%d' % (1200, 800, (master.winfo_screenwidth() - 1200) / 2, 0))
        # 1200*800 is size of window. screenwidth-1200/2 to center the window

        self.information = tkst.ScrolledText(self, wrap=tk.WORD, height=20, width=100)
        self.btn1 = tk.Button(self, text='Cours', command=self.CoursButton, height=3, width=40)
        self.btn2 = tk.Button(self, text='Absences', command=self.AbsButton, height=3, width=40)
        self.btn3 = tk.Button(self, text='Notes', command=self.NotesButton, height=3, width=40)
        self.btn5 = tk.Button(self, text='Update Cours', command=self.UpdateCours,height=3,width=40)
        self.btn6 = tk.Button(self, text='Update Absences', command=self.UpdateAbsences,height=3,width=40)
        self.btn7 = tk.Button(self, text='Update Notes', command=self.UpdateNotes,height=3,width=40)
        self.btn8 = tk.Button(self, text='Update All', command=self.UpdateAll,height=3,width=40)
        self.text_search = tk.Entry(self,width=40)

        self.information.config(font=font.Font(size=15))
        self.information.configure(background='#C0C0C0')

        self.btn2.config(font=font.Font(size=12))
        self.btn3.config(font=font.Font(size=12))
        self.btn1.config(font=font.Font(size=12))
        self.btn5.config(font=font.Font(size=12))
        self.btn6.config(font=font.Font(size=12))
        self.btn7.config(font=font.Font(size=12))
        self.btn8.config(font=font.Font(size=12))
        self.text_search.config(font=font.Font(size=12))

        self.information.insert(tk.INSERT, 'Click the boutons on the left side to check information from localhost file.\nClick the boutons on the right side to update information from website.\nAnd the area of text is for the search of cours with key word.\n')

        self.information.grid(row=0,column=0, columnspan=2)
        self.btn1.grid(row=1,column=0)
        self.btn2.grid(row=2,column=0)
        self.btn3.grid(row=3,column=0)
        self.text_search.grid(row=4,column=0)
        self.btn5.grid(row=1,column=1)
        self.btn6.grid(row=2,column=1)
        self.btn7.grid(row=3,column=1)
        self.btn8.grid(row=4,column=1)

        self.text_search.bind("<Return>",self.SearchCours)

        if(choice == 1):
            try:
                resabs = AnalyseAbs.analyse(inners=abs_download.connectEsv(writeFile=False, usr=usr, pwd=pwd))
            except:
                resabs = 'bug'
            self.information.delete("1.0", tk.END)
            if (resabs == 'unchanged'):
                self.information.insert(tk.INSERT, 'ABS isn\'t changed\n')
            elif (resabs == 'changed'):
                self.information.insert(tk.INSERT, 'New ABS has been saved\n')
            elif (resabs == 'bug'):
                self.information.insert(tk.INSERT, 'ABS bugged\n')
            elif (resabs == 'zero'):
                self.information.insert(tk.INSERT, 'You don\'t have any absence now\n')

            self.information.insert(tk.INSERT, '\n')
        elif(choice == 2):
            try:
                resnotes = AnalyseNotes.analyse(inners=notes_download.connectEsv(writeFile=False, usr=usr, pwd=pwd))
            except:
                resnotes = 'bug'
            self.information.delete("1.0", tk.END)
            if (resnotes == 'unchanged'):
                self.information.insert(tk.INSERT, 'Note isn\'t changed\n')
            elif (resnotes == 'changed'):
                self.information.insert(tk.INSERT, 'New note has been saved\n')
            elif (resnotes == 'bug'):
                self.information.insert(tk.INSERT, 'Notes bugged\n')

            self.information.insert(tk.INSERT, '\n')
        elif(choice == 3):
            try:
                rescours = AnalyseCours.analyse(
                    inners=cours_download.connectEsv(writeFile=False, usr=usr, pwd=pwd, nbWeeks=5))
            except:
                rescours = 'bug'
            self.information.delete("1.0", tk.END)
            if (rescours == 'unchanged'):
                self.information.insert(tk.INSERT, 'Cours are not changed or totally included by old one\n')
            elif (rescours == 'changed'):
                self.information.insert(tk.INSERT, 'New cours have been saved or some cours are changed\n')
            elif (rescours == 'bug'):
                self.information.insert(tk.INSERT, 'Cours bugged\n')
            self.information.insert(tk.INSERT,'\n')
        elif(choice == 4):
            try:
                resabs = AnalyseAbs.analyse(inners=abs_download.connectEsv(writeFile=False, usr=usr, pwd=pwd))
            except:
                resabs = 'bug'
            self.information.delete("1.0", tk.END)
            if (resabs == 'unchanged'):
                self.information.insert(tk.INSERT, 'ABS isn\'t changed\n')
            elif (resabs == 'changed'):
                self.information.insert(tk.INSERT, 'New ABS has been saved\n')
            elif (resabs == 'bug'):
                self.information.insert(tk.INSERT, 'ABS bugged\n')
            elif (resabs == 'zero'):
                self.information.insert(tk.INSERT, 'You don\'t have any absence now\n')

            try:
                resnotes = AnalyseNotes.analyse(inners=notes_download.connectEsv(writeFile=False, usr=usr, pwd=pwd))
            except:
                resnotes = 'bug'

            if (resnotes == 'unchanged'):
                self.information.insert(tk.INSERT, 'Note isn\'t changed\n')
            elif (resnotes == 'changed'):
                self.information.insert(tk.INSERT, 'New note has been saved\n')
            elif (resnotes == 'bug'):
                self.information.insert(tk.INSERT, 'Notes bugged\n')

            try:
                rescours = AnalyseCours.analyse(
                    inners=cours_download.connectEsv(writeFile=False, usr=usr, pwd=pwd, nbWeeks=5))
            except:
                rescours = 'bug'

            if (rescours == 'unchanged'):
                self.information.insert(tk.INSERT, 'Cours are not changed or totally included by old one\n')
            elif (rescours == 'changed'):
                self.information.insert(tk.INSERT, 'New cours have been saved or some cours are changed\n')
            elif (rescours == 'bug'):
                self.information.insert(tk.INSERT, 'Cours bugged\n')

            self.information.insert(tk.INSERT, '\n')

        choice = 0

    def CoursButton(self):
        readcours(self)

    def AbsButton(self):
        readabs(self)

    def NotesButton(self):
        readnotes(self)

    def UpdateCours(self):
        global choice
        choice = 3
        if(usr=='null' and pwd=='null'):
            self.master.change(LoginFrame)
        else:
            self.master.change(MainFrame)

    def UpdateAbsences(self):
        global choice
        choice = 1
        if(usr=='null' and pwd=='null'):
            self.master.change(LoginFrame)
        else:
            self.master.change(MainFrame)

    def UpdateNotes(self):
        global choice
        choice = 2
        if(usr=='null' and pwd=='null'):
            self.master.change(LoginFrame)
        else:
            self.master.change(MainFrame)

    def UpdateAll(self):
        global choice
        choice = 4
        if(usr=='null' and pwd=='null'):
            self.master.change(LoginFrame)
        else:
            self.master.change(MainFrame)

    def SearchCours(self,event):
        text = self.text_search.get()
        searchcours(text,self)

class LoginFrame(tk.Frame):
    def __init__(self,master=None, **kwargs):
        tk.Frame.__init__(self, master, **kwargs)
        master.title("Window to login")
        master.geometry('450x400')

        self.label_usr = tk.Label(self, height='2', text='Your username:')
        self.text_usr = tk.Entry(self, width=30)
        self.label_pwd = tk.Label(self,height='2', text='Your password:')
        self.text_pwd = tk.Entry(self,width=30, show='*')
        self.btn1 = tk.Button(self, text="Login and update information", command=self.LoginButton,height = 3,width = 30)
        self.btn2 = tk.Button(self, text='Return', command=self.ReturnButton,height = 3,width = 30)

        self.text_usr.bind("<Tab>", self.focus_next_window)
        self.text_pwd.bind("<Tab>", self.focus_next_window)
        self.text_pwd.bind("<Return>",self.LoginKey)

        self.text_usr.focus_set()

        self.btn2.config(font=font.Font(size=12))
        self.btn1.config(font=font.Font(size=12))
        self.label_usr.config(font=font.Font(size=12))
        self.label_pwd.config(font=font.Font(size=12))
        self.text_pwd.config(font=font.Font(size=18))
        self.text_usr.config(font=font.Font(size=18))

        self.label_usr.pack(pady=5)
        self.text_usr.pack(pady=5)
        self.label_pwd.pack(pady=5)
        self.text_pwd.pack(pady=5)
        self.btn1.pack(pady=5)
        self.btn2.pack(pady=5)

    def LoginButton(self):
        global usr
        global pwd
        usr = self.text_usr.get()
        pwd = self.text_pwd.get()
        self.master.change(MainFrame)

    def ReturnButton(self):
        global choice
        global usr
        global pwd
        choice = 0
        pwd = 'null'
        usr = 'null'
        self.master.change(MainFrame)

    def LoginKey(self,event):
        global usr
        global pwd
        usr = self.text_usr.get()
        pwd = self.text_pwd.get()
        self.master.change(MainFrame)

    def focus_next_window(self,event):
        event.widget.tk_focusNext().focus()
        return("break")

if __name__=="__main__":
    app=Mainframe()
    app.mainloop()